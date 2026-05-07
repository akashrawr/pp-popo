"""
lead_explainability.py
======================
Uses LLaMA3 (via Ollama) to generate a natural-language story for each lead,
grounded in SHAP values, topic alignment signals, and lead context.

Requirements:
    pip install pandas openpyxl requests tqdm

Ollama must be running locally:
    ollama serve
    ollama pull llama3
"""

import json
import time
import requests
import pandas as pd
from tqdm import tqdm

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
OLLAMA_URL        = "http://localhost:11434/api/generate"
MODEL             = "llama3"
INPUT_FILE        = "synthetic_leads_shap - copy.xlsx"
OUTPUT_FILE       = "leads_with_explanations.xlsx"
TOP_N_FEATURES    = 5      # how many SHAP drivers to highlight in the prompt
MAX_TOKENS        = 400    # cap on LLM response length
TEMPERATURE       = 0.4    # lower = more consistent/factual tone
BATCH_SIZE        = 10     # rows to process before saving a checkpoint
REQUEST_TIMEOUT   = 120    # seconds before giving up on one LLM call


# ─────────────────────────────────────────────
# STEP 1 — LOAD & PARSE
# ─────────────────────────────────────────────
def load_data(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    print(f"Loaded {len(df)} leads with {len(df.columns)} columns.")
    return df


def extract_shap_features(row: pd.Series, top_n: int = TOP_N_FEATURES) -> dict:
    """
    Pull all (feature_name, shap_value) pairs from a row, sort by |SHAP|,
    and return the top N positive drivers and top N negative drivers.
    """
    pairs = []
    i = 1
    while f"feature_name{i:02d}" in row.index:
        fname = row[f"feature_name{i:02d}"]
        fval  = row[f"shap_value{i:02d}"]
        if pd.notna(fname) and pd.notna(fval):
            pairs.append((fname, float(fval)))
        i += 1

    pairs.sort(key=lambda x: x[1], reverse=True)
    positive_drivers = [(f, v) for f, v in pairs if v > 0][:top_n]
    negative_drivers = [(f, v) for f, v in pairs if v < 0][-top_n:][::-1]  # most negative first

    return {
        "all_pairs":        pairs,
        "positive_drivers": positive_drivers,
        "negative_drivers": negative_drivers,
    }


def extract_topic_signals(row: pd.Series) -> list[dict]:
    """
    Parse topic alignment columns into a clean list of active/inactive signals.
    """
    topics = [
        ("Lead Description Completeness",    "Lead_Description_Completeness"),
        ("Lead Score & Stage",               "Lead_Score_and_Stage"),
        ("Corporate Email Address",          "Corporate_Email_Address"),
        ("Existing Customer Validity",       "Existing_Customer_Account_Validity"),
        ("Source Awareness",                 "Source_Awareness"),
        ("Early Intent Signal",              "Early_Intent"),
        ("Strong Commercial Intent",         "Strong_Commercial_Intent"),
        ("FSE Generated Lead",               "FSE_Generated"),
    ]
    signals = []
    for label, col in topics:
        alignment_col = f"{col}_alignment"
        shap_col      = f"{col}_shap"
        if alignment_col in row.index:
            signals.append({
                "label":     label,
                "aligned":   bool(row[alignment_col]),
                "shap":      float(row[shap_col]) if shap_col in row.index else 0.0,
            })
    return signals


# ─────────────────────────────────────────────
# STEP 2 — PROMPT ENGINEERING
# ─────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert aviation sales analyst. Your job is to write a clear, 
concise, and compelling lead story in ONE paragraph (4–6 sentences) for a sales representative.

Rules:
- Write in plain business English — no bullet points, no headers.
- Ground every claim in the data provided. Never invent information.
- Mention the lead rating, predicted conversion probability, and top drivers naturally in the narrative.
- Highlight what is working in the lead's favor AND what concerns exist.
- End with a recommended next action for the sales rep.
- Do NOT repeat the raw numbers verbatim — interpret them (e.g. "strong webinar engagement" instead of "shap_value = 1.35").
"""

def build_prompt(row: pd.Series, shap_info: dict, topic_signals: list[dict]) -> str:
    """
    Compose the full user prompt block from a lead row.
    """
    # ── Lead summary block ──
    lead_block = f"""
LEAD SUMMARY
------------
Lead ID       : {row['Lead_ID']}
Company       : {row['Company_Account']}
Lead Status   : {row['Lead_Status']}
Lead Source   : {row['Lead_Source']}
Region        : {row['Lead_Region']} / {row['Sub_Region']}
Market        : {row['Account_Market']}
Account Type  : {row['Account_Type']}
Product       : {row['Product_Interest']}
Aircraft      : {row['Aircraft_Type']}
Est. Value    : ${row['Estimated_Value_USD']:,.0f}
Loyalty       : {row['Loyalty_Segment'].capitalize()}
Account Exists: {row['Account_Exists']}
Phone         : {'Available' if row['Phone_Available'] else 'Not Available'}
Corporate Email: {'Yes' if row['Corporate_Email'] else 'No'}
""".strip()

    # ── Scores block ──
    score_block = f"""
MODEL SCORES
------------
Predicted Conversion Probability : {row['pred_proba']:.1%}
Lead Rating                      : {row['lead_rating'].upper()}
AQL Category                     : {row['aql_category'].replace('_', ' ').title()}
Behavior Score                   : {row['Behavior_Score']:.1f}/100
Demographic Score                : {row['Demographic_Score']:.1f}/100
Engagement Score                 : {row['Engagement_Score']:.1f}/100
Confidence                       : {row['pred_confidence']:.1%}
""".strip()

    # ── Top SHAP drivers ──
    pos_lines = "\n".join(
        f"  (+) {f}: strongly supports conversion" for f, _ in shap_info["positive_drivers"]
    )
    neg_lines = "\n".join(
        f"  (-) {f}: acts as a barrier to conversion" for f, _ in shap_info["negative_drivers"]
    )
    shap_block = f"""
KEY CONVERSION DRIVERS (from SHAP analysis)
--------------------------------------------
Positive Signals:
{pos_lines if pos_lines else '  None in top features'}

Negative Signals:
{neg_lines if neg_lines else '  None in top features'}
""".strip()

    # ── Topic alignment signals ──
    aligned     = [s["label"] for s in topic_signals if s["aligned"]]
    not_aligned = [s["label"] for s in topic_signals if not s["aligned"]]
    topic_block = f"""
TOPIC ALIGNMENT SIGNALS
-----------------------
Aligned    : {', '.join(aligned) if aligned else 'None'}
Not Aligned: {', '.join(not_aligned) if not_aligned else 'None'}
""".strip()

    # ── Final assembled prompt ──
    return f"""
Based on the following lead data, write a storytelling paragraph that explains this lead 
to a sales representative in a natural, insightful, and action-oriented way.

{lead_block}

{score_block}

{shap_block}

{topic_block}

Your narrative paragraph:
""".strip()


# ─────────────────────────────────────────────
# STEP 3 — LLM CALL
# ─────────────────────────────────────────────
def call_ollama(system: str, user_prompt: str) -> str:
    """
    Send a prompt to the local Ollama LLaMA3 instance and return the response text.
    """
    payload = {
        "model":  MODEL,
        "prompt": f"<|system|>\n{system}\n<|user|>\n{user_prompt}\n<|assistant|>",
        "stream": False,
        "options": {
            "temperature":   TEMPERATURE,
            "num_predict":   MAX_TOKENS,
            "stop":          ["<|user|>", "<|system|>"],
        },
    }
    try:
        resp = requests.post(OLLAMA_URL, json=payload, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        return data.get("response", "").strip()
    except requests.exceptions.ConnectionError:
        return "[ERROR] Cannot connect to Ollama. Is 'ollama serve' running?"
    except requests.exceptions.Timeout:
        return "[ERROR] Request timed out. Try reducing MAX_TOKENS or simplifying the prompt."
    except Exception as e:
        return f"[ERROR] {str(e)}"


# ─────────────────────────────────────────────
# STEP 4 — PROCESS ALL LEADS
# ─────────────────────────────────────────────
def process_leads(df: pd.DataFrame) -> pd.DataFrame:
    """
    Iterate over every lead row, build the prompt, call LLaMA3, and store the result.
    """
    explanations  = []
    prompts_used  = []
    durations     = []

    print(f"\nRunning LLaMA3 explanations for {len(df)} leads...\n")

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Generating"):
        shap_info     = extract_shap_features(row)
        topic_signals = extract_topic_signals(row)
        prompt        = build_prompt(row, shap_info, topic_signals)

        t0       = time.time()
        response = call_ollama(SYSTEM_PROMPT, prompt)
        elapsed  = round(time.time() - t0, 2)

        explanations.append(response)
        prompts_used.append(prompt)
        durations.append(elapsed)

        # checkpoint save every BATCH_SIZE rows
        if (idx + 1) % BATCH_SIZE == 0:
            _checkpoint_save(df, explanations, prompts_used, durations, idx)

    df["llm_explanation"]  = explanations
    df["prompt_used"]      = prompts_used
    df["llm_duration_sec"] = durations
    return df


def _checkpoint_save(df, explanations, prompts_used, durations, up_to_idx):
    """Save a partial checkpoint so no work is lost on long runs."""
    temp = df.iloc[:len(explanations)].copy()
    temp["llm_explanation"]  = explanations
    temp["prompt_used"]      = prompts_used
    temp["llm_duration_sec"] = durations
    ckpt = OUTPUT_FILE.replace(".xlsx", f"_checkpoint_{up_to_idx+1}.xlsx")
    temp.to_excel(ckpt, index=False)
    tqdm.write(f"  ✓ Checkpoint saved → {ckpt}")


# ─────────────────────────────────────────────
# STEP 5 — SAVE OUTPUT
# ─────────────────────────────────────────────
def save_output(df: pd.DataFrame, path: str):
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl import load_workbook

    df.to_excel(path, index=False, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active

    # Style header row
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Auto-width + wrap explanation column
    for col in ws.columns:
        col_letter = col[0].column_letter
        header_val = ws[f"{col_letter}1"].value or ""
        if "explanation" in str(header_val).lower():
            ws.column_dimensions[col_letter].width = 80
            for cell in col[1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        elif "prompt" in str(header_val).lower():
            ws.column_dimensions[col_letter].width = 60
        else:
            ws.column_dimensions[col_letter].width = 18

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(path)
    print(f"\n✅ Final output saved → {path}")


# ─────────────────────────────────────────────
# STEP 6 — DEMO / DRY-RUN (no Ollama needed)
# ─────────────────────────────────────────────
def demo_prompt(df: pd.DataFrame, row_index: int = 0):
    """
    Print the fully assembled prompt for one lead without calling Ollama.
    Useful for iterating on prompt design.
    """
    row           = df.iloc[row_index]
    shap_info     = extract_shap_features(row)
    topic_signals = extract_topic_signals(row)
    prompt        = build_prompt(row, shap_info, topic_signals)

    print("=" * 70)
    print(f"SYSTEM PROMPT:\n{SYSTEM_PROMPT}")
    print("=" * 70)
    print(f"USER PROMPT (Lead index {row_index} — {row['Lead_ID']}):\n")
    print(prompt)
    print("=" * 70)
    return prompt


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="LLaMA3 Lead Explainability Engine")
    parser.add_argument("--demo",   action="store_true",
                        help="Print prompt for lead 0 without calling Ollama")
    parser.add_argument("--row",    type=int, default=0,
                        help="Row index to use in --demo mode (default: 0)")
    parser.add_argument("--limit",  type=int, default=None,
                        help="Process only the first N leads (useful for testing)")
    parser.add_argument("--input",  type=str, default=INPUT_FILE,
                        help=f"Input Excel file (default: {INPUT_FILE})")
    parser.add_argument("--output", type=str, default=OUTPUT_FILE,
                        help=f"Output Excel file (default: {OUTPUT_FILE})")
    args = parser.parse_args()

    df = load_data(args.input)

    if args.demo:
        # ── Just preview the prompt, don't call LLM ──
        demo_prompt(df, row_index=args.row)

    else:
        # ── Full run ──
        if args.limit:
            print(f"⚠  Limiting to first {args.limit} leads.")
            df = df.head(args.limit).copy()

        df = process_leads(df)
        save_output(df, args.output)

        # Quick summary stats
        errors = df["llm_explanation"].str.startswith("[ERROR]").sum()
        avg_t  = df["llm_duration_sec"].mean()
        print(f"\nSummary:")
        print(f"  Total leads processed : {len(df)}")
        print(f"  Errors                : {errors}")
        print(f"  Avg. time per lead    : {avg_t:.1f}s")
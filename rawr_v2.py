"""
Enterprise-grade Lead Explanation Generator
-------------------------------------------
Improved version with:
- Proper SHAP ranking by absolute magnitude
- Deterministic interpretation layer
- Hallucination guardrails
- Structured evidence generation
- Confidence-aware language
- Reduced prompt verbosity
- Explanation validation
- Stable / reproducible outputs
"""

import re
import json
import time
import requests
import pandas as pd
from tqdm import tqdm
from typing import Dict, List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# CONFIG
OLLAMA_URL      = "http://localhost:11434/api/generate"
MODEL           = "llama3"
INPUT_FILE      = "synthetic_leads_shap - copy.xlsx"
OUTPUT_FILE     = "leads_with_explanations.xlsx"
TOP_N_FEATURES  = 5
MAX_TOKENS      = 250
TEMPERATURE     = 0.15
BATCH_SIZE      = 10
REQUEST_TIMEOUT = 120

# FEATURE INTERPRETATION LAYER
# This converts raw ML features into deterministic
# business-safe interpretations BEFORE the LLM sees them.
# This is the key enterprise-grade improvement.

FEATURE_INTERPRETATIONS = {

    "Behavior_Score": {
        "positive": "The lead demonstrates strong behavioral engagement.",
        "negative": "The lead shows limited behavioral engagement."
    },

    "Engagement_Score": {
        "positive": "The lead has actively engaged with marketing activity.",
        "negative": "Engagement activity has been relatively low."
    },

    "Demographic_Score": {
        "positive": "The lead aligns well with the target customer profile.",
        "negative": "The lead has weaker alignment with the target customer profile."
    },

    "Corporate_Email": {
        "positive": "A verified corporate email is available.",
        "negative": "The lead does not use a verified corporate email."
    },

    "Phone_Available": {
        "positive": "Direct contact information is available.",
        "negative": "Direct contact information is limited."
    },

    "Account_Exists": {
        "positive": "An existing customer account relationship exists.",
        "negative": "No existing customer relationship was identified."
    },

    "Lead_Source": {
        "positive": "The lead originated from a high-quality acquisition source.",
        "negative": "The acquisition source has historically shown weaker conversion performance."
    }
}

# VALIDATION GUARDRAILS
FORBIDDEN_PATTERNS = [
    r"\bbudget\b",
    r"\bdecision-maker\b",
    r"\bprocurement\b",
    r"\burgent\b",
    r"\bhigh intent\b",
    r"\bguaranteed\b",
    r"\bwill convert\b",
    r"\bready to buy\b",
]

# SYSTEM PROMPT
SYSTEM_PROMPT = """
You are an AI sales explanation assistant.

Your task is to generate a factual, concise business explanation
for why a lead was qualified or disqualified.

STRICT RULES:
- Use ONLY the supplied evidence.
- Do NOT invent facts.
- Do NOT infer:
  - budget
  - authority
  - urgency
  - technical requirements
  - procurement readiness
- Do NOT exaggerate certainty.
- Keep the tone professional and analytical.
- Output EXACTLY one paragraph.
- Keep the explanation between 4 and 6 sentences.
- Mention:
  - predicted conversion likelihood
  - strongest positive drivers
  - strongest negative drivers
  - recommended sales action
"""

# LOAD DATA
def load_data(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    print(f"Loaded {len(df)} leads.")
    return df

# SHAP EXTRACTION (FIXED)
def extract_shap_features(
    row: pd.Series,
    top_n: int = TOP_N_FEATURES
) -> Dict:

    pairs = []

    i = 1
    while f"feature_name{i:02d}" in row.index:

        fname = row[f"feature_name{i:02d}"]
        sval  = row[f"shap_value{i:02d}"]

        if pd.notna(fname) and pd.notna(sval):
            pairs.append((fname, float(sval)))

        i += 1

    # Sort by ABSOLUTE SHAP magnitude
    pairs_sorted = sorted(
        pairs,
        key=lambda x: abs(x[1]),
        reverse=True
    )

    positive_drivers = [
        (f, v)
        for f, v in pairs_sorted
        if v > 0
    ][:top_n]

    negative_drivers = [
        (f, v)
        for f, v in pairs_sorted
        if v < 0
    ][:top_n]

    return {
        "all_pairs": pairs_sorted,
        "positive_drivers": positive_drivers,
        "negative_drivers": negative_drivers,
    }

# CONFIDENCE LANGUAGE
def confidence_phrase(prob: float) -> str:
    if prob >= 0.90:
        return "strongly indicates"
    elif prob >= 0.75:
        return "suggests"
    elif prob >= 0.60:
        return "shows moderate signals of"
    else:
        return "shows limited evidence of"

# DETERMINISTIC INTERPRETATION LAYER
def interpret_driver(feature_name: str, shap_value: float) -> str:
    direction = "positive" if shap_value > 0 else "negative"
    if feature_name in FEATURE_INTERPRETATIONS:
        return FEATURE_INTERPRETATIONS[feature_name][direction]
    # Generic fallback
    if shap_value > 0:
        return f"{feature_name} contributed positively to qualification likelihood."

    return f"{feature_name} reduced qualification likelihood."

# TOPIC SIGNALS
def extract_topic_signals(row: pd.Series):
    topics = [
        ("Lead Description Completeness", "Lead_Description_Completeness"),
        ("Lead Score & Stage", "Lead_Score_and_Stage"),
        ("Corporate Email Address", "Corporate_Email_Address"),
        ("Existing Customer Validity", "Existing_Customer_Account_Validity"),
        ("Source Awareness", "Source_Awareness"),
        ("Early Intent Signal", "Early_Intent"),
        ("Strong Commercial Intent", "Strong_Commercial_Intent"),
        ("FSE Generated Lead", "FSE_Generated"),
    ]

    signals = []
    for label, col in topics:

        alignment_col = f"{col}_alignment"

        if alignment_col in row.index:

            signals.append({
                "label": label,
                "aligned": bool(row[alignment_col])
            })

    return signals

# STRUCTURED EVIDENCE GENERATION
def build_structured_evidence(
    row,
    shap_info,
    topic_signals
):

    positives = [
        {
            "feature": f,
            "shap": round(v, 4),
            "meaning": interpret_driver(f, v)
        }
        for f, v in shap_info["positive_drivers"]
    ]

    negatives = [
        {
            "feature": f,
            "shap": round(v, 4),
            "meaning": interpret_driver(f, v)
        }
        for f, v in shap_info["negative_drivers"]
    ]

    aligned_topics = [
        s["label"]
        for s in topic_signals
        if s["aligned"]
    ]

    not_aligned_topics = [
        s["label"]
        for s in topic_signals
        if not s["aligned"]
    ]

    evidence = {

        "lead_summary": {

            "lead_id": row["Lead_ID"],
            "company": row["Company_Account"],
            "lead_rating": row["lead_rating"],
            "predicted_probability": round(row["pred_proba"], 4),
            "confidence_level": confidence_phrase(row["pred_proba"]),
            "product_interest": row["Product_Interest"],
            "market": row["Account_Market"],
        },

        "positive_evidence": positives,
        "negative_evidence": negatives,

        "aligned_topics": aligned_topics,
        "not_aligned_topics": not_aligned_topics,

        "recommended_action": (
            "Prioritize timely follow-up and validate engagement interest."
            if row["pred_proba"] >= 0.70
            else
            "Conduct additional qualification before prioritizing sales outreach."
        )
    }

    return evidence

# PROMPT BUILDER
def build_prompt(evidence: Dict) -> str:

    return f"""
Generate a factual lead explanation using ONLY the evidence below.

EVIDENCE:
{json.dumps(evidence, indent=2)}

Write one concise business paragraph.
""".strip()

# OLLAMA CALL
def call_ollama(system: str, user_prompt: str) -> str:
    payload = {
        "model": MODEL,
        "prompt":
            f"<|system|>\n{system}\n"
            f"<|user|>\n{user_prompt}\n"
            f"<|assistant|>",
        "stream": False,
        "options": {
            "temperature": TEMPERATURE,
            "num_predict": MAX_TOKENS,
            "stop": ["<|user|>", "<|system|>"],
        },
    }
    try:
        resp = requests.post(
            OLLAMA_URL,
            json=payload,
            timeout=REQUEST_TIMEOUT
        )
        resp.raise_for_status()
        data = resp.json()
        return data.get("response", "").strip()
    except Exception as e:
        return f"[ERROR] {str(e)}"

# VALIDATION
def validate_explanation(text: str) -> bool:
    text_lower = text.lower()
    for pattern in FORBIDDEN_PATTERNS:
        if re.search(pattern, text_lower):
            return False
    return True

# PROCESS LEADS
def process_leads(df: pd.DataFrame):
    explanations = []
    prompts_used = []
    durations = []
    validation_flags = []
    print(f"\nGenerating explanations for {len(df)} leads...\n")

    for idx, row in tqdm(
        df.iterrows(),
        total=len(df),
        desc="Generating"
    ):
        shap_info = extract_shap_features(row)
        topic_signals = extract_topic_signals(row)
        evidence = build_structured_evidence(
            row,
            shap_info,
            topic_signals
        )
        prompt = build_prompt(evidence)
        t0 = time.time()
        response = call_ollama(
            SYSTEM_PROMPT,
            prompt
        )
        elapsed = round(time.time() - t0, 2)

        # VALIDATION CHECK
        valid = validate_explanation(response)

        if not valid:
            response = (
                "[FLAGGED] Explanation contained unsupported claims."
            )

        explanations.append(response)
        prompts_used.append(prompt)
        durations.append(elapsed)
        validation_flags.append(valid)

        # checkpoint
        if (idx + 1) % BATCH_SIZE == 0:
            checkpoint = df.iloc[:len(explanations)].copy()
            checkpoint["llm_explanation"] = explanations
            checkpoint["validation_passed"] = validation_flags
            ckpt_name = OUTPUT_FILE.replace(
                ".xlsx",
                f"_checkpoint_{idx+1}.xlsx"
            )
            checkpoint.to_excel(ckpt_name, index=False)
            tqdm.write(f"✓ Checkpoint saved → {ckpt_name}")

    df["llm_explanation"] = explanations
    df["prompt_used"] = prompts_used
    df["llm_duration_sec"] = durations
    df["validation_passed"] = validation_flags

    return df

# SAVE OUTPUT
def save_output(df: pd.DataFrame, path: str):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    header_fill = PatternFill(
        "solid",
        start_color="1F4E79"
    )
    header_font = Font(
        bold=True,
        color="FFFFFF"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True
        )

    for col in ws.columns:
        col_letter = col[0].column_letter
        header_val = str(
            ws[f"{col_letter}1"].value
        ).lower()
        if "explanation" in header_val:
            ws.column_dimensions[col_letter].width = 80
            for cell in col[1:]:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )
        elif "prompt" in header_val:
            ws.column_dimensions[col_letter].width = 60
        else:
            ws.column_dimensions[col_letter].width = 20
    ws.freeze_panes = "A2"
    wb.save(path)

    print(f"\nSaved final output → {path}")

if __name__ == "__main__":
    df = load_data(INPUT_FILE)
    final_df = process_leads(df)
    save_output(final_df, OUTPUT_FILE)
    print("\nDone.")